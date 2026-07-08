"""Excel exporter for analytics reports (Phase 11.7)."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.analytics.services.report_payload import AnalyticsReportPayload, format_report_value


class ExcelReportExporter:
    """Export analytics reports as multi-sheet Excel workbooks."""

    _HEADER_FONT = Font(bold=True)

    def export(self, payload: AnalyticsReportPayload) -> bytes:
        """Render *payload* as an XLSX workbook."""
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Overview"
        self._write_overview_sheet(overview, payload)

        metrics = workbook.create_sheet("Metrics")
        self._write_metrics_sheet(metrics, payload)

        trends = workbook.create_sheet("Trend Data")
        self._write_trends_sheet(trends, payload)

        tables = workbook.create_sheet("Raw Tables")
        self._write_tables_sheet(tables, payload)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _write_overview_sheet(self, sheet, payload: AnalyticsReportPayload) -> None:
        rows = [
            ("Report Title", payload.module_title),
            ("Module", payload.module),
            ("Generated At", format_report_value(payload.generated_at)),
            ("Start Date", format_report_value(payload.start_date)),
            ("End Date", format_report_value(payload.end_date)),
        ]
        for row_index, (label, value) in enumerate(rows, start=1):
            sheet.cell(row=row_index, column=1, value=label).font = self._HEADER_FONT
            sheet.cell(row=row_index, column=2, value=value)
        self._autosize_columns(sheet, max_column=2)

    def _write_metrics_sheet(self, sheet, payload: AnalyticsReportPayload) -> None:
        sheet.cell(row=1, column=1, value="Metric").font = self._HEADER_FONT
        sheet.cell(row=1, column=2, value="Value").font = self._HEADER_FONT
        for row_index, (label, value) in enumerate(payload.kpis, start=2):
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=value)
        self._autosize_columns(sheet, max_column=2)

    def _write_trends_sheet(self, sheet, payload: AnalyticsReportPayload) -> None:
        row_index = 1
        for trend in payload.trends:
            sheet.cell(row=row_index, column=1, value=trend.name).font = self._HEADER_FONT
            row_index += 1
            sheet.cell(row=row_index, column=1, value="Date").font = self._HEADER_FONT
            sheet.cell(row=row_index, column=2, value="Value").font = self._HEADER_FONT
            row_index += 1
            for date_key, value in sorted(trend.points.items()):
                sheet.cell(row=row_index, column=1, value=date_key)
                sheet.cell(row=row_index, column=2, value=format_report_value(value))
                row_index += 1
            row_index += 1
        self._autosize_columns(sheet, max_column=2)

    def _write_tables_sheet(self, sheet, payload: AnalyticsReportPayload) -> None:
        row_index = 1
        for table in payload.tables:
            sheet.cell(row=row_index, column=1, value=table.name).font = self._HEADER_FONT
            row_index += 1
            if not table.headers:
                row_index += 1
                continue
            for column_index, header in enumerate(table.headers, start=1):
                sheet.cell(row=row_index, column=column_index, value=header).font = self._HEADER_FONT
            row_index += 1
            for table_row in table.rows:
                for column_index, value in enumerate(table_row, start=1):
                    sheet.cell(row=row_index, column=column_index, value=value)
                row_index += 1
            row_index += 1
        max_column = max((len(table.headers) for table in payload.tables if table.headers), default=1)
        self._autosize_columns(sheet, max_column=max_column)

    @staticmethod
    def _autosize_columns(sheet, *, max_column: int) -> None:
        for column_index in range(1, max_column + 1):
            column_letter = get_column_letter(column_index)
            max_length = 0
            for cell in sheet[column_letter]:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)
